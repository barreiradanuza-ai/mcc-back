"""
Seed script: populates coverage tables from final.xlsx.

Usage:
    DATABASE_URL=postgres://user:pass@host:5432/db python seed_db.py [path/to/final.xlsx]

The spreadsheet must have the following sheets:
  - CLARO     : one CEP per row (integer, zero-padded to 8 digits)
  - TIM       : header "TIM" in row 1, then one CEP per row
  - PROMOCLARO: header "CIDADES" in row 1, then one city name per row
"""

import os
import sys
import unicodedata

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)

import psycopg2
import psycopg2.extras

DATABASE_URL = os.getenv("DATABASE_URL", "")


def normalize_city(text: str) -> str:
    """Uppercase + strip accents, matching the lookup in scraper.py."""
    return unicodedata.normalize("NFD", text.upper()).encode("ascii", "ignore").decode().strip()


def pad_cep(value) -> str | None:
    """Convert an integer or string CEP to an 8-digit zero-padded string."""
    if value is None:
        return None
    try:
        digits = str(int(value)).zfill(8)
        if len(digits) != 8:
            return None
        return digits
    except (ValueError, TypeError):
        return None


def load_ceps(ws, skip_header: bool = False) -> list[tuple[str]]:
    """Extract valid 8-digit CEPs from a worksheet column A."""
    ceps = []
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
        if skip_header and i == 0:
            continue
        cep = pad_cep(row[0])
        if cep:
            ceps.append((cep,))
    return ceps


def load_cities(ws) -> list[tuple[str]]:
    """Extract normalized city names from a worksheet column A (skips header row)."""
    cities = []
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
        if i == 0:
            continue
        if row[0] and str(row[0]).strip():
            cities.append((normalize_city(str(row[0])),))
    return cities


def seed(xlsx_path: str):
    if not DATABASE_URL:
        print("Error: DATABASE_URL environment variable is not set.")
        sys.exit(1)

    print(f"Opening workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    print("Reading CLARO sheet...")
    ceps_claro = load_ceps(wb["CLARO"], skip_header=False)
    print(f"  {len(ceps_claro):,} CEPs loaded")

    print("Reading TIM sheet...")
    ceps_tim = load_ceps(wb["TIM"], skip_header=True)
    print(f"  {len(ceps_tim):,} CEPs loaded")

    print("Reading PROMOCLARO sheet...")
    cities_promo = load_cities(wb["PROMOCLARO"])
    print(f"  {len(cities_promo):,} cities loaded")

    wb.close()

    print("\nConnecting to Postgres...")
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            print("Creating tables if not exist...")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ceps_claro (
                    cep CHAR(8) PRIMARY KEY
                );
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ceps_tim (
                    cep CHAR(8) PRIMARY KEY
                );
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cidades_promo_claro (
                    cidade TEXT PRIMARY KEY
                );
            """)

            print("Truncating existing data...")
            cur.execute("TRUNCATE TABLE ceps_claro")
            cur.execute("TRUNCATE TABLE ceps_tim")
            cur.execute("TRUNCATE TABLE cidades_promo_claro")

            print(f"Inserting {len(ceps_claro):,} Claro CEPs...")
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO ceps_claro (cep) VALUES %s ON CONFLICT DO NOTHING",
                ceps_claro,
                page_size=5000,
            )

            print(f"Inserting {len(ceps_tim):,} TIM CEPs...")
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO ceps_tim (cep) VALUES %s ON CONFLICT DO NOTHING",
                ceps_tim,
                page_size=5000,
            )

            print(f"Inserting {len(cities_promo):,} Claro Promo cities...")
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO cidades_promo_claro (cidade) VALUES %s ON CONFLICT DO NOTHING",
                cities_promo,
            )

        conn.commit()
        print("\nSeed completed successfully.")
    except Exception as e:
        conn.rollback()
        print(f"Error during seed: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "final.xlsx"
    seed(path)
